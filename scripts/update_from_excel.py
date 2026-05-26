from __future__ import annotations

import json
import re
import subprocess
import unicodedata
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import range_boundaries
from zipfile import ZipFile
from xml.etree import ElementTree as ET


ROOT = Path(__file__).resolve().parents[1]
INDEX = ROOT / "index.html"
DOCS = ROOT / "docs"

AREA_KEY = {
    "국 어": "korean",
    "수 학": "math",
    "영 어": "english",
    "사 회": "society",
    "과 학": "science",
    "체 육": "pe",
    "예 술": "arts",
    "기술": "tech",
    "제2외국어": "second",
    "교 양": "career",
}

PREFIX = {
    "korean": "k",
    "math": "m",
    "english": "e",
    "history": "h",
    "society": "s",
    "science": "sc",
    "pe": "p",
    "arts": "a",
    "tech": "t",
    "second": "l",
    "career": "c",
}

TYPE_MAP = {"공통": "il", "일반": "il", "진로": "jin", "융합": "fus"}

FIXED_IDS = {
    "공통국어1": "k1",
    "공통국어2": "k2",
    "문학": "k3",
    "화법과 언어": "k4",
    "매체 의사소통": "k5",
    "독서 토론과 글쓰기": "k6",
    "독서와 작문": "k7",
    "언어생활 탐구": "k8",
    "주제 탐구 독서": "k9",
    "문학과 영상": "k10",
    "공통수학1": "m1",
    "공통수학2": "m2",
    "대수": "m3",
    "확률과 통계": "m4",
    "미적분Ⅰ": "m5",
    "기하": "m6",
    "미적분Ⅱ": "m7",
    "인공지능 수학": "m8",
    "경제 수학": "m9",
    "수학과제 탐구": "m10",
    "수학과 문화": "m11",
    "실용 통계": "m12",
    "고급 대수": "m13",
    "공통영어1": "e1",
    "공통영어2": "e2",
    "영어Ⅰ": "e3",
    "영어Ⅱ": "e4",
    "영어 독해와 작문": "e5",
    "영미 문학 읽기": "e6",
    "영어 발표와 토론": "e7",
    "심화 영어 독해와 작문": "e8",
    "한국사1": "h1",
    "한국사2": "h2",
    "통합사회1": "s1",
    "통합사회2": "s2",
    "현대사회와 윤리": "s3",
    "세계시민과 지리": "s4",
    "사회와 문화": "s5",
    "세계사": "s6",
    "경제": "s7",
    "윤리와 사상": "s8",
    "도시의 미래 탐구": "s9",
    "한국지리 탐구": "s10",
    "인문학과 윤리": "s11",
    "사회문제 탐구": "s12",
    "역사로 탐구하는 현대 세계": "s13",
    "여행지리": "s14",
    "금융과 경제생활": "s15",
    "윤리문제 탐구": "s16",
    "통합과학1": "sc1",
    "통합과학2": "sc2",
    "과학탐구실험1": "sc3",
    "과학탐구실험2": "sc4",
    "물리학": "sc5",
    "화학": "sc6",
    "생명과학": "sc7",
    "지구과학": "sc8",
    "역학과 에너지": "sc9",
    "물질과 에너지": "sc10",
    "세포와 물질대사": "sc11",
    "행성우주과학": "sc12",
    "전자기와 양자": "sc13",
    "화학 반응의 세계": "sc14",
    "생물의 유전": "sc15",
    "지구시스템과학": "sc16",
    "기후변화와 환경생태": "sc17",
    "융합과학 탐구": "sc18",
    "고급 물리학": "sc19",
    "고급 화학": "sc20",
    "고급 생명과학": "sc21",
    "고급 지구과학": "sc22",
    "체육1": "p1",
    "체육2": "p2",
    "스포츠 문화": "p3",
    "스포츠 과학": "p4",
    "스포츠 생활1": "p5",
    "스포츠 생활2": "p6",
    "음악": "a1",
    "미술": "a2",
    "음악 연주와 창작": "a3",
    "미술 창작": "a4",
    "음악 감상과 비평": "a5",
    "미술과 매체": "a6",
    "정보": "t1",
    "인공지능 기초": "t2",
    "컴퓨터 시스템 일반": "t3",
    "데이터 과학": "t4",
    "소프트웨어와 생활": "t5",
    "정보과학": "t6",
    "일본어": "l1",
    "중국어": "l2",
    "일본어 회화": "l3",
    "중국어 회화": "l4",
    "일본 문화": "l5",
    "중국 문화": "l6",
    "한문": "l7",
    "진로와 직업": "c1",
    "논술": "c2",
}

COURSE_ALIASES = {
    "영어독해와 작문": ["영어 독해와 작문"],
    "생물과 유전": ["생물의 유전"],
    "확률": ["확률과 통계"],
    "미적분": ["미적분Ⅰ", "미적분Ⅱ"],
}

AREA_SLASH_PREFIXES = ("제2외국어",)

MATCH_KEYWORDS = {
    "국어국문": ["국어국문"],
    "사학": ["사학"],
    "철학": ["철학"],
    "심리": ["심리"],
    "교육": ["교육"],
    "사회": ["사회학"],
    "언론": ["언론"],
    "경제": ["경제"],
    "경영": ["경영"],
    "법학": ["법학"],
    "행정": ["행정"],
    "국제관계": ["국제관계"],
    "사회복지": ["사회복지"],
    "수학": ["수학"],
    "물리": ["물리"],
    "화학": ["화학"],
    "생명과학": ["생명과학"],
    "컴퓨터": ["컴퓨터"],
    "전기전자": ["전기전자", "전자전기"],
    "기계": ["기계"],
    "화공": ["화공", "화학공학"],
    "건축": ["건축"],
    "의예": ["의예", "의학"],
    "치의예": ["치의예", "치의학"],
    "한의예": ["한의예", "한의학"],
    "약학": ["약학"],
    "간호": ["간호"],
    "수의예": ["수의예", "수의학"],
    "국어교육": ["국어교육"],
    "수학교육": ["수학교육"],
    "영어교육": ["영어교육"],
    "과학교육": ["과학교육"],
    "사회교육": ["사회교육"],
    "체육교육": ["체육교육"],
    "미술": ["미술"],
    "음악": ["음악"],
    "체육": ["체육"],
    "디자인": ["디자인"],
    "무용": ["무용"],
    "농학": ["농학", "농업"],
    "식품영양": ["식품영양"],
    "산림": ["산림"],
    "환경생태": ["환경생태", "생태환경"],
}

GENERIC_TEXT = (
    "진로 및 적성",
    "전 과목",
    "교과목 선택",
)


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def norm_name(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def get_area(raw: str) -> str | None:
    for key, area in AREA_KEY.items():
        if key in raw:
            return area
    return None


def build_course_to_id() -> dict[str, str]:
    subjects = []
    for year in ("2026", "2025"):
        subjects.extend(item for values in load_curriculum(year).values() for item in values)
    course_to_id: dict[str, str] = {}
    for item in subjects:
        course_to_id.setdefault(item["name"], item["id"])
    for name, subject_id in FIXED_IDS.items():
        course_to_id.setdefault(name, subject_id)
    return course_to_id


def curriculum_path_for_year(year: str) -> Path:
    grade = "1학년" if year == "2026" else "2학년"
    for path in sorted(DOCS.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        name = norm_name(path.name)
        if year in name and norm_name(grade) in name and norm_name("입학") in name:
            return path
    raise FileNotFoundError(f"배당표 xlsx not found for {year} in {DOCS}")


def curriculum_to_dict(subjects: OrderedDict[str, list[dict[str, object]]]) -> dict[str, list[dict[str, object]]]:
    return {area: list(items) for area, items in subjects.items()}


def load_curriculum_from_path(path: Path) -> OrderedDict[str, list[dict[str, object]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next(wb[name] for name in wb.sheetnames if "입학생" in name)
    values = load_curriculum_values(ws)
    subjects: OrderedDict[str, list[dict[str, object]]] = OrderedDict(
        (key, []) for key in PREFIX
    )

    for row in range(7, 106):
        area = get_area(clean(values.get((row, 2))))
        name = clean(values.get((row, 4)))
        typ = clean(values.get((row, 3)))
        if not name or typ not in TYPE_MAP or not area:
            continue

        subject_area = "history" if name.startswith("한국사") else area
        for col in range(7, 13):
            if values.get((row, col)) is None:
                continue
            subjects[subject_area].append(
                {
                    "id": FIXED_IDS[name],
                    "name": name,
                    "type": TYPE_MAP[typ],
                    "grade": (col - 7) // 2 + 1,
                    "sem": (col - 7) % 2 + 1,
                }
            )

    return subjects


def load_curriculum(year: str) -> OrderedDict[str, list[dict[str, object]]]:
    return load_curriculum_from_path(curriculum_path_for_year(year))


def load_curriculum_values(ws) -> dict[tuple[int, int], object]:
    values = {}
    for row in range(1, ws.max_row + 1):
        for col in (2, 3, 4, 7, 8, 9, 10, 11, 12):
            values[(row, col)] = ws.cell(row, col).value

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if not set(range(min_col, max_col + 1)) & {2, 3, 4, 7, 8, 9, 10, 11, 12}:
            continue
        source = values.get((min_row, min_col))
        if source is None:
            continue
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if col in {2, 3, 4, 7, 8, 9, 10, 11, 12} and values.get((row, col)) is None:
                    values[(row, col)] = source
    return values


def js_value(value: object, indent: int = 2) -> str:
    sp = " " * indent
    if isinstance(value, dict):
        parts = []
        for key, val in value.items():
            parts.append(f"{sp}{json.dumps(key, ensure_ascii=False)}: {js_value(val, indent + 2)}")
        return "{\n" + ",\n".join(parts) + "\n" + " " * (indent - 2) + "}"
    if isinstance(value, list):
        if not value:
            return "[]"
        if isinstance(value[0], dict) and "id" in value[0]:
            lines = []
            for item in value:
                fields = ", ".join(
                    [
                        f"id:'{item['id']}'",
                        f"name:'{item['name']}'",
                        f"type:'{item['type']}'",
                        f"grade:{item['grade']}",
                        f"sem:{item['sem']}",
                    ]
                )
                lines.append(f"{sp}{{ {fields} }}")
            return "[\n" + ",\n".join(lines) + "\n" + " " * (indent - 2) + "]"
        if isinstance(value[0], dict):
            return "[" + ",".join(js_value(item, indent + 2) for item in value) + "]"
        return "[" + ",".join(f"'{item}'" for item in value) + "]"
    if isinstance(value, str):
        return json.dumps(value, ensure_ascii=False)
    return json.dumps(value, ensure_ascii=False)


def load_depts() -> dict[str, dict[str, object]]:
    script = """
const fs=require('fs');
const html=fs.readFileSync('index.html','utf8');
const m=html.match(/const DEPT_DATA = ([\\s\\S]*?);\\s*\\nconst MATCH_KEYWORDS/);
if(!m) throw new Error('DEPT_DATA block not found');
console.log(JSON.stringify(Function('return '+m[1])().depts));
"""
    return json.loads(subprocess.check_output(["node", "-e", script], cwd=ROOT, text=True))


def normalize_course_text(text: object) -> list[str]:
    raw = clean(text)
    if not raw or raw == "-" or any(marker in raw for marker in GENERIC_TEXT):
        return []
    for prefix in AREA_SLASH_PREFIXES:
        if raw.startswith(f"{prefix}/"):
            # F·G열 "제2외국어/한문" 등은 교과군 안내이지 구체 과목 권장이 아님
            return []
    raw = raw.replace("Ⅰ", "I").replace("Ⅱ", "II")
    raw = raw.replace("미적분II", "미적분Ⅱ").replace("미적분I", "미적분Ⅰ")
    raw = re.sub(r"[\[\]\(\)]", ",", raw)
    raw = re.sub(r"또는|및|/|·|\\n|-일반선택:|-진로선택:", ",", raw)
    tokens = []
    for token in raw.split(","):
        token = clean(token).strip(": ")
        if not token or token in {"국어", "영어", "수학", "사회", "과학", "기술", "가정"}:
            continue
        if token in COURSE_ALIASES:
            tokens.extend(COURSE_ALIASES[token])
            continue
        tokens.append(token)
    return tokens


def matches(unit: str, dept: str) -> bool:
    keywords = MATCH_KEYWORDS[dept]
    return any(keyword in unit for keyword in keywords)


def parse_unit_columns(col_d: object, col_e: object) -> tuple[str, str]:
    """D/E 분리 시 D=단과대·계열, E=모집단위. 병합·단일 셀은 unit만."""
    d = clean(col_d)
    e = clean(col_e)
    if e and d and e != d:
        return d, e
    return "", e or d


def university_key(name: object) -> str:
    raw = clean(name)
    if "서울대" in raw:
        return "seoul"
    if "연세대" in raw:
        return "yonsei"
    if "고려대" in raw:
        return "korea"
    for key in ("성균관", "한양", "이화", "중앙", "경희", "건국", "동국", "국민"):
        if key in raw:
            return key
    return raw


def ordered_unique(values: list[str]) -> list[str]:
    return list(OrderedDict((value, None) for value in values))


def row_subject_ids(row: dict[str, object], course_to_id: dict[str, str]) -> tuple[list[str], list[str], list[str], str]:
    row_core_ids: list[str] = []
    row_rec_ids: list[str] = []
    row_ref_ids: list[str] = []
    for token in normalize_course_text(row["core"]):
        row_core_ids.extend(token_to_ids(token, course_to_id))
    target = row_core_ids if row["rec_merged"] else row_rec_ids
    for token in normalize_course_text(row["rec"]):
        target.extend(token_to_ids(token, course_to_id))
    for token in normalize_course_text(row.get("ref_detail")):
        row_ref_ids.extend(token_to_ids(token, course_to_id))
    row_core = ordered_unique(row_core_ids)
    row_rec = [item for item in ordered_unique(row_rec_ids) if item not in set(row_core)]
    row_ref = ordered_unique(row_ref_ids)
    return row_core, row_rec, row_ref, clean(row["comment"])


def build_univ_index(rows: list[dict[str, object]], course_to_id: dict[str, str]) -> dict[str, dict[str, object]]:
    index: dict[str, dict[str, object]] = {}
    for row in rows:
        univ_name = clean(row["university"])
        unit = clean(row["unit"])
        if not univ_name or not unit:
            continue
        row_core, row_rec, row_ref, comment = row_subject_ids(row, course_to_id)
        if not (row_core or row_rec or row_ref or comment):
            continue
        univ = university_key(row["university"])
        unit_group = clean(row.get("unit_group"))
        entry = {
            "unitGroup": unit_group,
            "unit": unit,
            "comment": comment,
            "core": row_core,
            "rec": row_rec,
            "ref": row_ref,
        }
        bucket = index.setdefault(univ, {"univName": univ_name, "groups": {}})
        groups: dict[str, list[dict[str, object]]] = bucket["groups"]
        group_list = groups.setdefault(unit_group, [])
        existing = next((item for item in group_list if item["unit"] == unit), None)
        if existing:
            existing["core"] = ordered_unique(existing["core"] + entry["core"])
            existing["rec"] = ordered_unique(existing["rec"] + entry["rec"])
            existing["ref"] = ordered_unique(existing["ref"] + entry["ref"])
            if entry["comment"]:
                prev = clean(existing["comment"])
                existing["comment"] = " ".join(ordered_unique([prev, entry["comment"]])[:2])
        else:
            group_list.append(entry)
    return index


def build_global_index(rows: list[dict[str, object]], course_to_id: dict[str, str]) -> dict[str, object]:
    """전국: 모집단위명(E열) 기준으로 모든 대학 행을 합침."""
    by_unit: dict[str, dict[str, object]] = {}
    for row in rows:
        univ_name = clean(row["university"])
        unit = clean(row["unit"])
        if not univ_name or not unit:
            continue
        row_core, row_rec, row_ref, comment = row_subject_ids(row, course_to_id)
        if not (row_core or row_rec or row_ref or comment):
            continue
        unit_group = clean(row.get("unit_group"))
        source = {
            "univ": university_key(row["university"]),
            "univName": univ_name,
            "unitGroup": unit_group,
            "unit": unit,
            "comment": comment,
            "core": row_core,
            "rec": row_rec,
            "ref": row_ref,
        }
        if unit not in by_unit:
            by_unit[unit] = {
                "unitGroup": "",
                "unit": unit,
                "comment": comment,
                "core": list(row_core),
                "rec": list(row_rec),
                "ref": list(row_ref),
                "sources": [source],
            }
            continue
        entry = by_unit[unit]
        entry["core"] = ordered_unique(entry["core"] + row_core)
        entry["rec"] = ordered_unique(entry["rec"] + row_rec)
        entry["ref"] = ordered_unique(entry["ref"] + row_ref)
        if comment:
            prev = clean(entry["comment"])
            entry["comment"] = " ".join(ordered_unique([prev, comment])[:2])
        entry["sources"].append(source)
    units = sorted(by_unit.values(), key=lambda item: item["unit"])
    return {"univName": "전국 · 모집단위", "groups": {"": units}}


def load_dept_mapping(
    course_to_id: dict[str, str],
    depts: dict[str, dict[str, object]],
    path: Path | None = None,
    rows: list[dict[str, object]] | None = None,
) -> dict[str, dict[str, object]]:
    if rows is None:
        path = path or next(p for p in DOCS.glob("*.xlsx") if "2028" in p.name)
        rows = load_recommendation_rows(path)
    result = {}

    for dept, info in depts.items():
        core_ids: list[str] = []
        rec_ids: list[str] = []
        ref_ids: list[str] = []
        comments: list[str] = []
        sources = []
        for row in rows:
            unit = clean(row["unit"])
            if not matches(unit, dept):
                continue
            row_core, row_rec, row_ref, comment = row_subject_ids(row, course_to_id)
            if comment:
                comments.append(comment)
            core_ids.extend(row_core)
            rec_ids.extend(row_rec)
            ref_ids.extend(row_ref)
            if row_core or row_rec or row_ref or comment:
                sources.append(
                    {
                        "univ": university_key(row["university"]),
                        "univName": clean(row["university"]),
                        "unitGroup": clean(row.get("unit_group")),
                        "unit": unit,
                        "comment": comment,
                        "core": row_core,
                        "rec": row_rec,
                        "ref": row_ref,
                    }
                )
        core = ordered_unique(core_ids)
        rec = [item for item in ordered_unique(rec_ids) if item not in set(core)]
        ref = ordered_unique(ref_ids)
        result[dept] = {
            "comment": " ".join(ordered_unique(comments)[:2]),
            "core": core,
            "rec": rec,
            "ref": ref,
            "sources": sources,
        }
    return result


def token_to_ids(token: str, course_to_id: dict[str, str]) -> list[str]:
    subject_id = course_to_id.get(token)
    if subject_id:
        return [subject_id]
    matched = []
    for course_name, course_id in course_to_id.items():
        if course_name in token:
            matched.append(course_id)
    return matched


def load_recommendation_rows(path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    values = {}
    merged_recommendation_cells = set()
    for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), 5):
        for col in (3, 4, 5, 6, 7, 8, 9, 10):
            values[(row_num, col)] = row[col - 1] if len(row) >= col else None

    merge_ranges = read_sheet1_merge_ranges(path)
    for ref in merge_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if not set(range(min_col, max_col + 1)) & {3, 4, 5, 6, 7, 8, 9, 10}:
            continue
        source = values.get((min_row, min_col))
        if source is None:
            continue
        for row_num in range(max(min_row, 5), max_row + 1):
            for col in range(min_col, max_col + 1):
                if col in {6, 7}:
                    merged_recommendation_cells.add((row_num, col))
                if col == 9:
                    merged_recommendation_cells.add((row_num, col))
                if col in {3, 4, 5, 6, 7, 8, 9, 10} and values.get((row_num, col)) is None:
                    values[(row_num, col)] = source

    rows = []
    for row_num in range(5, ws.max_row + 1):
        unit_group, unit = parse_unit_columns(values.get((row_num, 4)), values.get((row_num, 5)))
        rows.append(
            {
                "university": values.get((row_num, 3)),
                "unit_group": unit_group,
                "unit": unit,
                "core": values.get((row_num, 6)),
                "rec": values.get((row_num, 7)),
                "ref": values.get((row_num, 8)),
                "ref_detail": values.get((row_num, 9)),
                "comment": values.get((row_num, 10)),
                "core_merged": (row_num, 6) in merged_recommendation_cells,
                "rec_merged": (row_num, 7) in merged_recommendation_cells,
                "ref_merged": (row_num, 9) in merged_recommendation_cells,
            }
        )
    return rows


def read_sheet1_merge_ranges(path: Path) -> list[str]:
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as archive:
        root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    return [node.attrib["ref"] for node in root.findall(".//m:mergeCell", ns)]


def format_subjects(name: str, subjects: OrderedDict[str, list[dict[str, object]]]) -> str:
    lines = [f"const {name} = {{"]
    for area, items in subjects.items():
        lines.append(f"  {area}: [")
        for item in items:
            lines.append(
                "    { "
                f"id:'{item['id']}', name:'{item['name']}', type:'{item['type']}', "
                f"grade:{item['grade']}, sem:{item['sem']} "
                "},"
            )
        lines.append("  ],")
    lines.append("};")
    return "\n".join(lines)


def format_depts(depts: dict[str, dict[str, object]]) -> str:
    lines = ["  depts: {"]
    for dept, info in depts.items():
        lines.append(
            f"    '{dept}': {{ comment:{json.dumps(info['comment'], ensure_ascii=False)}, "
            f"core:{js_value(info['core'])}, rec:{js_value(info['rec'])}, ref:{js_value(info['ref'])}, sources:{js_value(info['sources'])} }},"
        )
    lines.append("  },")
    return "\n".join(lines)


def main() -> None:
    html = INDEX.read_text(encoding="utf-8")
    subjects_2026 = load_curriculum("2026")
    subjects_2025 = load_curriculum("2025")
    course_to_id = build_course_to_id()
    excel_path = next(p for p in DOCS.glob("*.xlsx") if "2028" in p.name)
    excel_rows = load_recommendation_rows(excel_path)
    univ_index = build_univ_index(excel_rows, course_to_id)
    univ_index["__global__"] = build_global_index(excel_rows, course_to_id)

    subjects_block = (
        format_subjects("SUBJECTS_2026", subjects_2026)
        + "\n\n/* 2025 입학생 — 2025학년도 입학생 편제 */\n"
        + format_subjects("SUBJECTS_2025", subjects_2025)
        + "\n\n"
    )
    if "const SUBJECTS_2026 = {" in html:
        html = re.sub(
            r"const SUBJECTS_2026 = \{[\s\S]*?const SUBJECTS_2025 = \{[\s\S]*?\};\n\n",
            subjects_block,
            html,
            count=1,
        )
    elif "const AREAS = {" in html:
        html = re.sub(
            r"(const AREAS = \{[\s\S]*?\};\n\n)",
            r"\1" + subjects_block,
            html,
            count=1,
        )
    html = html.replace(
        'id="cell-${s.id}" data-id="${s.id}"',
        'data-id="${s.id}"',
    )
    html = html.replace(
        "const el = document.getElementById('cell-'+id);\n    if (el) el.classList.add('hl-core');",
        "document.querySelectorAll(`[data-id=\"${id}\"]`).forEach(el => el.classList.add('hl-core'));",
    )
    html = html.replace(
        "const el = document.getElementById('cell-'+id);\n    if (el) { el.classList.remove('hl-core'); el.classList.add('hl-rec'); }",
        "document.querySelectorAll(`[data-id=\"${id}\"]`).forEach(el => { el.classList.remove('hl-core'); el.classList.add('hl-rec'); });",
    )
    html = html.replace(
        "const el = document.getElementById('cell-'+id);\n    if (el && !el.classList.contains('hl-core') && !el.classList.contains('hl-rec')) {\n      el.classList.add('hl-ref');\n    }",
        "document.querySelectorAll(`[data-id=\"${id}\"]`).forEach(el => {\n      if (!el.classList.contains('hl-core') && !el.classList.contains('hl-rec')) el.classList.add('hl-ref');\n    });",
    )
    INDEX.write_text(html, encoding="utf-8")

    data_dir = ROOT / "data"
    data_dir.mkdir(exist_ok=True)
    payload = {
        "updatedAt": __import__("datetime").datetime.now(__import__("datetime").timezone.utc)
        .astimezone()
        .strftime("%Y-%m-%d %H:%M:%S %Z"),
        "source": next(p.name for p in DOCS.glob("*.xlsx") if "2028" in p.name),
        "unitCount": len(univ_index["__global__"]["groups"].get("", [])),
        "univIndex": univ_index,
        "subjects": {
            "2026": curriculum_to_dict(subjects_2026),
            "2025": curriculum_to_dict(subjects_2025),
        },
    }
    (data_dir / "depts.json").write_text(
        json.dumps(payload, ensure_ascii=False), encoding="utf-8"
    )

    print(f"subjects_2026={sum(len(values) for values in subjects_2026.values())}")
    print(f"subjects_2025={sum(len(values) for values in subjects_2025.values())}")
    print(f"univIndex={len(univ_index)}")
    print(f"globalUnits={len(univ_index['__global__']['groups'].get('', []))}")


if __name__ == "__main__":
    main()
